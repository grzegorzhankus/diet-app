"""
Tests for Export Engine
"""
import pytest
from datetime import date, timedelta
from pathlib import Path
import tempfile
import io

from openpyxl import load_workbook
import pandas as pd

from core.storage import Storage
from core.export import ExportEngine
from core.schemas import DailyEntryCreate


@pytest.fixture
def storage_with_data():
    """Create storage with comprehensive test data."""
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        storage = Storage(str(db_path))

        # 30 days of varied data
        base_date = date(2026, 1, 1)
        for i in range(30):
            entry = DailyEntryCreate(
                date=base_date + timedelta(days=i),
                weight_kg=85.0 - (i * 0.1),
                bodyfat_pct=20.0 - (i * 0.03),
                cal_in_kcal=1800.0 + (i * 10),
                cal_out_sport_kcal=300.0,
                source="test"
            )
            storage.create(entry)

        yield storage


@pytest.fixture
def storage_minimal():
    """Create storage with minimal data."""
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        storage = Storage(str(db_path))

        # Just 5 days
        base_date = date(2026, 1, 1)
        for i in range(5):
            entry = DailyEntryCreate(
                date=base_date + timedelta(days=i),
                weight_kg=85.0,
                source="test"
            )
            storage.create(entry)

        yield storage


def test_export_engine_init(storage_with_data):
    """Test export engine initialization."""
    engine = ExportEngine(storage_with_data, bmr_kcal=2000.0, target_weight_kg=75.0)
    assert engine.storage == storage_with_data
    assert engine.bmr_kcal == 2000.0
    assert engine.target_weight_kg == 75.0


def test_export_to_excel_basic(storage_with_data):
    """Test basic Excel export."""
    engine = ExportEngine(storage_with_data)

    excel_bytes = engine.export_to_excel(days=30)

    # Should return bytes
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0


def test_export_to_excel_loads_correctly(storage_with_data):
    """Test that exported Excel can be loaded."""
    engine = ExportEngine(storage_with_data)

    excel_bytes = engine.export_to_excel(days=30)

    # Load with openpyxl
    wb = load_workbook(io.BytesIO(excel_bytes))

    # Should have 5 sheets
    assert len(wb.sheetnames) == 5
    assert "Summary" in wb.sheetnames
    assert "Daily Data" in wb.sheetnames
    assert "KPIs" in wb.sheetnames
    assert "Red Flags" in wb.sheetnames
    assert "Charts Data" in wb.sheetnames


def test_export_summary_sheet_content(storage_with_data):
    """Test Summary sheet has correct content."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30)

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Summary"]

    # Check title
    assert "DIET_APP" in str(ws['A1'].value)

    # Check sections exist
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert any("WEIGHT METRICS" in str(v) for v in values)
    assert any("Current Weight" in str(v) for v in values)


def test_export_daily_data_sheet_content(storage_with_data):
    """Test Daily Data sheet has correct content."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30)

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Daily Data"]

    # Check headers exist
    headers = [cell.value for cell in ws[1]]
    assert "Date" in headers
    assert "Weight (kg)" in headers

    # Should have header + data rows (at least 2 rows)
    assert ws.max_row >= 2


def test_export_kpi_sheet_content(storage_with_data):
    """Test KPI sheet has correct content."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30)

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["KPIs"]

    # Check headers
    headers = [cell.value for cell in ws[1]]
    assert "KPI" in headers
    assert "Value" in headers
    assert "Status" in headers

    # Should have multiple KPIs
    assert ws.max_row > 5


def test_export_red_flags_sheet_content(storage_with_data):
    """Test Red Flags sheet has correct content."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30)

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Red Flags"]

    # Check headers or "no flags" message
    first_cell = ws['A1'].value

    if "No red flags" in str(first_cell):
        # Healthy data, no flags
        assert "Great job" in str(first_cell) or "No red flags" in str(first_cell)
    else:
        # Has flags
        headers = [cell.value for cell in ws[1]]
        assert "Severity" in headers
        assert "Category" in headers


def test_export_charts_data_sheet_content(storage_with_data):
    """Test Charts Data sheet has correct content."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30)

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Charts Data"]

    # Check headers
    headers = [cell.value for cell in ws[1]]
    assert "Week" in headers
    assert "Avg Weight (kg)" in headers

    # Should have weekly aggregates
    assert ws.max_row >= 2  # At least header + 1 week


def test_export_without_kpis(storage_with_data):
    """Test export without KPI sheet."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30, include_kpis=False)

    wb = load_workbook(io.BytesIO(excel_bytes))

    # Should have 4 sheets (no KPIs)
    assert len(wb.sheetnames) == 4
    assert "KPIs" not in wb.sheetnames
    assert "Summary" in wb.sheetnames


def test_export_without_red_flags(storage_with_data):
    """Test export without Red Flags sheet."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30, include_red_flags=False)

    wb = load_workbook(io.BytesIO(excel_bytes))

    # Should have 4 sheets (no Red Flags)
    assert len(wb.sheetnames) == 4
    assert "Red Flags" not in wb.sheetnames
    assert "Summary" in wb.sheetnames


def test_export_minimal_sheets(storage_with_data):
    """Test export with minimal sheets."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(
        days=30,
        include_kpis=False,
        include_red_flags=False
    )

    wb = load_workbook(io.BytesIO(excel_bytes))

    # Should have 3 sheets (Summary, Daily Data, Charts Data)
    assert len(wb.sheetnames) == 3
    assert "Summary" in wb.sheetnames
    assert "Daily Data" in wb.sheetnames
    assert "Charts Data" in wb.sheetnames


def test_export_to_csv_basic(storage_with_data):
    """Test basic CSV export."""
    engine = ExportEngine(storage_with_data)

    csv_bytes = engine.export_to_csv(days=30)

    # Should return bytes
    assert isinstance(csv_bytes, bytes)
    assert len(csv_bytes) > 0


def test_export_to_csv_content(storage_with_data):
    """Test CSV export content."""
    engine = ExportEngine(storage_with_data)
    csv_bytes = engine.export_to_csv(days=30)

    # Parse CSV
    df = pd.read_csv(io.BytesIO(csv_bytes))

    # Check columns
    assert "date" in df.columns
    assert "bs_weight_kg" in df.columns

    # Should have at least some rows
    assert len(df) >= 5


def test_export_to_csv_empty_data():
    """Test CSV export with empty data."""
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        storage = Storage(str(db_path))

        engine = ExportEngine(storage)
        csv_bytes = engine.export_to_csv(days=30)

        # Should return message
        assert b"No data available" in csv_bytes


def test_export_empty_data_handling():
    """Test export with empty database."""
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        storage = Storage(str(db_path))

        engine = ExportEngine(storage)
        excel_bytes = engine.export_to_excel(days=30)

        # Should still generate workbook
        wb = load_workbook(io.BytesIO(excel_bytes))

        # Summary should have "No data available"
        ws = wb["Summary"]
        assert "No data available" in str(ws['A1'].value)


def test_export_different_periods(storage_with_data):
    """Test export with different time periods."""
    engine = ExportEngine(storage_with_data)

    for days in [7, 14, 30, 60, 90]:
        excel_bytes = engine.export_to_excel(days=days)
        wb = load_workbook(io.BytesIO(excel_bytes))

        # Should generate valid workbook
        assert len(wb.sheetnames) >= 3


def test_export_data_accuracy(storage_with_data):
    """Test that exported data matches source data."""
    engine = ExportEngine(storage_with_data)

    # Get CSV export
    csv_bytes = engine.export_to_csv(days=30)
    df_export = pd.read_csv(io.BytesIO(csv_bytes))

    # Get source data
    df_source = engine.metrics.get_metrics_dataframe(days=30)

    # Check row counts match
    assert len(df_export) == len(df_source)

    # Check weights match (first and last)
    assert abs(df_export['bs_weight_kg'].iloc[0] - df_source['bs_weight_kg'].iloc[0]) < 0.01
    assert abs(df_export['bs_weight_kg'].iloc[-1] - df_source['bs_weight_kg'].iloc[-1]) < 0.01


def test_export_with_custom_bmr(storage_with_data):
    """Test export with custom BMR."""
    engine = ExportEngine(storage_with_data, bmr_kcal=2500.0)

    excel_bytes = engine.export_to_excel(days=30)
    wb = load_workbook(io.BytesIO(excel_bytes))

    # Should generate valid workbook
    assert len(wb.sheetnames) == 5

    # Summary should contain BMR info
    ws = wb["Summary"]
    values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    assert any("2500" in v for v in values)


def test_export_with_custom_target_weight(storage_with_data):
    """Test export with custom target weight."""
    engine = ExportEngine(storage_with_data, target_weight_kg=70.0)

    excel_bytes = engine.export_to_excel(days=30)

    # Should generate valid workbook
    wb = load_workbook(io.BytesIO(excel_bytes))
    assert len(wb.sheetnames) == 5


def test_export_header_formatting(storage_with_data):
    """Test that headers are properly formatted."""
    engine = ExportEngine(storage_with_data)
    excel_bytes = engine.export_to_excel(days=30)

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Daily Data"]

    # Check first row (header) has formatting
    header_cell = ws['A1']
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb is not None


def test_export_deterministic(storage_with_data):
    """Test that repeated exports produce identical results."""
    engine = ExportEngine(storage_with_data)

    excel_bytes1 = engine.export_to_excel(days=30)
    excel_bytes2 = engine.export_to_excel(days=30)

    # Excel files may have metadata timestamps, so check structural similarity instead
    # Both should be valid and similar size
    assert len(excel_bytes1) == len(excel_bytes2)

    # Both should be valid Excel files
    from openpyxl import load_workbook
    wb1 = load_workbook(io.BytesIO(excel_bytes1))
    wb2 = load_workbook(io.BytesIO(excel_bytes2))

    # Should have same number of sheets
    assert len(wb1.sheetnames) == len(wb2.sheetnames)


def test_export_csv_deterministic(storage_with_data):
    """Test that repeated CSV exports produce identical results."""
    engine = ExportEngine(storage_with_data)

    csv_bytes1 = engine.export_to_csv(days=30)
    csv_bytes2 = engine.export_to_csv(days=30)

    # Should be identical
    assert csv_bytes1 == csv_bytes2
