"""
Export Engine for DIET_APP
Generates Excel and CSV exports with comprehensive analytics.
"""
from datetime import date, timedelta
from typing import Optional, Dict, List
from pathlib import Path
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from core.storage import Storage
from core.metrics import MetricsEngine
from core.kpi_engine import KPIEngine
from core.red_flags import RedFlagsEngine


class ExportEngine:
    """Generates comprehensive exports in Excel and CSV formats."""

    def __init__(self, storage: Storage, bmr_kcal: float = 2000.0, target_weight_kg: float = 75.0):
        """
        Initialize export engine.

        Args:
            storage: Storage instance
            bmr_kcal: Basal Metabolic Rate in kcal/day
            target_weight_kg: Target weight for KPI calculations
        """
        self.storage = storage
        self.bmr_kcal = bmr_kcal
        self.target_weight_kg = target_weight_kg

        self.metrics = MetricsEngine(storage, bmr_kcal=bmr_kcal)
        self.kpis = KPIEngine(storage, target_weight_kg=target_weight_kg, bmr_kcal=bmr_kcal)
        self.red_flags = RedFlagsEngine(storage, bmr_kcal=bmr_kcal)

    def export_to_excel(
        self,
        days: int = 90,
        include_kpis: bool = True,
        include_red_flags: bool = True
    ) -> bytes:
        """
        Export comprehensive report to Excel format.

        Args:
            days: Number of days to include
            include_kpis: Include KPI sheet
            include_red_flags: Include red flags sheet

        Returns:
            Excel file as bytes
        """
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Summary
        self._create_summary_sheet(wb, days)

        # Sheet 2: Daily Data
        self._create_daily_data_sheet(wb, days)

        # Sheet 3: KPIs (if requested)
        if include_kpis:
            self._create_kpi_sheet(wb, days)

        # Sheet 4: Red Flags (if requested)
        if include_red_flags:
            self._create_red_flags_sheet(wb, days)

        # Sheet 5: Charts Data (aggregated metrics for easy charting)
        self._create_charts_data_sheet(wb, days)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_summary_sheet(self, wb: Workbook, days: int):
        """Create summary statistics sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Get stats
        stats = self.metrics.get_summary_stats(days=days)

        if not stats:
            ws['A1'] = "No data available"
            return

        # Title
        ws['A1'] = "DIET_APP - Summary Report"
        ws['A1'].font = Font(size=16, bold=True)

        # Metadata
        row = 3
        ws[f'A{row}'] = "Report Date:"
        ws[f'B{row}'] = date.today().strftime('%Y-%m-%d')
        row += 1
        ws[f'A{row}'] = "Period:"
        ws[f'B{row}'] = f"Last {days} days"
        row += 1
        ws[f'A{row}'] = "Data Coverage:"
        ws[f'B{row}'] = f"{stats.get('data_coverage', 0):.1f}%"
        row += 2

        # Weight metrics
        ws[f'A{row}'] = "WEIGHT METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metrics_data = [
            ("Current Weight", f"{stats.get('weight_current_kg', 0):.1f} kg"),
            ("Start Weight", f"{stats.get('weight_start_kg', 0):.1f} kg"),
            ("Weight Change", f"{stats.get('weight_change_kg', 0):+.1f} kg"),
            ("Average Weight", f"{stats.get('weight_avg_kg', 0):.1f} kg"),
            ("Min Weight", f"{stats.get('weight_min_kg', 0):.1f} kg"),
            ("Max Weight", f"{stats.get('weight_max_kg', 0):.1f} kg"),
            ("7-Day Average", f"{stats.get('weight_7d_avg_kg', 0):.1f} kg" if stats.get('weight_7d_avg_kg') else "N/A"),
            ("14-Day Average", f"{stats.get('weight_14d_avg_kg', 0):.1f} kg" if stats.get('weight_14d_avg_kg') else "N/A"),
        ]

        for label, value in metrics_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Body composition
        if stats.get('fm_current_kg'):
            ws[f'A{row}'] = "BODY COMPOSITION"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            comp_data = [
                ("Current Fat Mass", f"{stats.get('fm_current_kg', 0):.1f} kg"),
                ("Average Fat Mass", f"{stats.get('fm_avg_kg', 0):.1f} kg"),
                ("Current Body Fat %", f"{stats.get('bf_current_pct', 0):.1f}%"),
                ("Average Body Fat %", f"{stats.get('bf_avg_pct', 0):.1f}%"),
            ]

            for label, value in comp_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

            row += 1

        # Calorie metrics
        if stats.get('cal_net_avg_kcal') is not None:
            ws[f'A{row}'] = "CALORIE METRICS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            cal_data = [
                ("Avg NET Calories", f"{stats.get('cal_net_avg_kcal', 0):.0f} kcal/day"),
                ("Avg Intake", f"{stats.get('cal_in_avg_kcal', 0):.0f} kcal/day"),
                ("Avg Sport", f"{stats.get('cal_out_sport_avg_kcal', 0):.0f} kcal/day"),
                ("BMR", f"{stats.get('cal_out_bmr_kcal', 0):.0f} kcal/day"),
                ("Avg Total Output", f"{stats.get('cal_out_total_avg_kcal', 0):.0f} kcal/day"),
            ]

            for label, value in cal_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_daily_data_sheet(self, wb: Workbook, days: int):
        """Create daily data sheet with all metrics."""
        ws = wb.create_sheet("Daily Data")

        # Get metrics
        df = self.metrics.get_metrics_dataframe(days=days)

        if df.empty:
            ws['A1'] = "No data available"
            return

        # Prepare columns
        columns = [
            'date',
            'bs_weight_kg',
            'bs_bodyfat_pct',
            'bs_fat_mass_kg',
            'bs_lean_mass_kg',
            'cal_in_kcal',
            'cal_out_sport_kcal',
            'cal_net_kcal',
            'bs_weight_7d_avg_kg',
            'bs_weight_14d_avg_kg',
            'cal_net_7d_avg_kcal'
        ]

        # Filter to available columns
        available_cols = [col for col in columns if col in df.columns]
        df_export = df[available_cols].copy()

        # Rename for better readability
        rename_map = {
            'date': 'Date',
            'bs_weight_kg': 'Weight (kg)',
            'bs_bodyfat_pct': 'Body Fat (%)',
            'bs_fat_mass_kg': 'Fat Mass (kg)',
            'bs_lean_mass_kg': 'Lean Mass (kg)',
            'cal_in_kcal': 'Calories IN',
            'cal_out_sport_kcal': 'Calories SPORT',
            'cal_net_kcal': 'Calories NET',
            'bs_weight_7d_avg_kg': 'Weight 7d Avg',
            'bs_weight_14d_avg_kg': 'Weight 14d Avg',
            'cal_net_7d_avg_kcal': 'NET 7d Avg'
        }

        df_export = df_export.rename(columns=rename_map)

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Header formatting
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width

    def _create_kpi_sheet(self, wb: Workbook, days: int):
        """Create KPI sheet."""
        ws = wb.create_sheet("KPIs")

        # Get KPIs
        kpis = self.kpis.compute_all_kpis(days=days)

        if not kpis:
            ws['A1'] = "No KPIs available"
            return

        # Headers
        headers = ['KPI', 'Value', 'Unit', 'Status', 'Window', 'Explanation']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Data
        for r_idx, kpi in enumerate(kpis, 2):
            ws.cell(row=r_idx, column=1, value=kpi.name)
            ws.cell(row=r_idx, column=2, value=kpi.value if kpi.value is not None else "N/A")
            ws.cell(row=r_idx, column=3, value=kpi.unit)

            # Status
            status_cell = ws.cell(row=r_idx, column=4)
            if kpi.is_good is True:
                status_cell.value = "✓ Good"
                status_cell.font = Font(color="00AA00")
            elif kpi.is_good is False:
                status_cell.value = "✗ Needs Work"
                status_cell.font = Font(color="AA0000")
            else:
                status_cell.value = "-"

            ws.cell(row=r_idx, column=5, value=f"{kpi.window_days}d" if kpi.window_days else "-")
            ws.cell(row=r_idx, column=6, value=kpi.explanation)

        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 60

    def _create_red_flags_sheet(self, wb: Workbook, days: int):
        """Create red flags sheet."""
        ws = wb.create_sheet("Red Flags")

        # Get red flags
        flags = self.red_flags.detect_all_flags(days=days)

        if not flags:
            ws['A1'] = "No red flags detected - Great job!"
            ws['A1'].font = Font(color="00AA00", bold=True)
            return

        # Headers
        headers = ['Severity', 'Category', 'Issue', 'Description', 'Recommendation']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Data
        severity_colors = {
            'critical': 'FF0000',
            'high': 'FF6600',
            'medium': 'FFAA00',
            'low': 'FFFF00'
        }

        for r_idx, flag in enumerate(flags, 2):
            # Severity
            sev_cell = ws.cell(row=r_idx, column=1, value=flag.severity.upper())
            sev_cell.fill = PatternFill(start_color=severity_colors.get(flag.severity, 'FFFFFF'),
                                       end_color=severity_colors.get(flag.severity, 'FFFFFF'),
                                       fill_type="solid")
            sev_cell.font = Font(bold=True)

            ws.cell(row=r_idx, column=2, value=flag.category)
            ws.cell(row=r_idx, column=3, value=flag.title)
            ws.cell(row=r_idx, column=4, value=flag.description)
            ws.cell(row=r_idx, column=5, value=flag.recommendation or "")

        # Auto-size columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 60

    def _create_charts_data_sheet(self, wb: Workbook, days: int):
        """Create aggregated data for easy charting."""
        ws = wb.create_sheet("Charts Data")

        # Get weekly aggregates
        df = self.metrics.get_metrics_dataframe(days=days)

        if df.empty:
            ws['A1'] = "No data available"
            return

        # Convert date to datetime and add week
        df['date'] = pd.to_datetime(df['date'])
        df['week'] = df['date'].dt.to_period('W').astype(str)

        # Weekly aggregates
        weekly = df.groupby('week').agg({
            'bs_weight_kg': 'mean',
            'bs_fat_mass_kg': 'mean',
            'cal_in_kcal': 'mean',
            'cal_out_sport_kcal': 'mean',
            'cal_net_kcal': 'mean'
        }).reset_index()

        weekly.columns = ['Week', 'Avg Weight (kg)', 'Avg Fat Mass (kg)',
                         'Avg Calories IN', 'Avg Calories SPORT', 'Avg Calories NET']

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(weekly, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Header formatting
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width

    def export_to_csv(self, days: int = 90) -> bytes:
        """
        Export daily data to CSV format.

        Args:
            days: Number of days to include

        Returns:
            CSV file as bytes
        """
        # Get metrics
        df = self.metrics.get_metrics_dataframe(days=days)

        if df.empty:
            return b"No data available"

        # Prepare columns
        columns = [
            'date',
            'bs_weight_kg',
            'bs_bodyfat_pct',
            'bs_fat_mass_kg',
            'bs_lean_mass_kg',
            'cal_in_kcal',
            'cal_out_sport_kcal',
            'cal_net_kcal',
            'bs_weight_7d_avg_kg',
            'bs_weight_14d_avg_kg'
        ]

        # Filter to available columns
        available_cols = [col for col in columns if col in df.columns]
        df_export = df[available_cols].copy()

        # Convert to CSV
        output = io.StringIO()
        df_export.to_csv(output, index=False)

        return output.getvalue().encode('utf-8')
